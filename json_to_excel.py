"""
VTU JSON to Excel Aggregation Module
This module aggregates all JSON mark extraction results into a structured Excel file.

Process:
1. Load all JSON files from gemini_json_results folder
2. Canonicalize subject codes (handle variations like BCS405A vs BCS405A.)
3. Filter subjects based on user-provided subject codes (if any)
4. Create/update Excel file with structured format:
   - One row per USN
   - Columns grouped by subject (Internal, External, Total, Result)
5. Apply color coding: Green (P), Red (F), Yellow (A)
"""

import os
import re
import shutil
import json
import difflib
import sys 
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

# ==============================================
# HELPER FUNCTION: Normalize Subject Codes
# ==============================================
def normalize_code(code):
    """
    Normalize subject code by removing non-alphanumeric characters and converting to uppercase.
    This helps match variations like "BCS405A", "BCS405A.", "bcs405a" to the same code.
    
    Args:
        code: Subject code string (may contain special characters)
        
    Returns:
        Normalized code string (uppercase, alphanumeric only)
    """
    if not code: return ""
    return re.sub(r"[^A-Z0-9]", "", str(code).upper())


# ==============================================
# CONFIGURATION
# ==============================================
# Folder containing JSON files (one per USN screenshot)
JSON_INPUT_FOLDER = "gemini_json_results"

# Output Excel filename
EXCEL_OUTPUT_FILENAME = "vtu_structured_results.xlsx"

# Hybrid Canonicalization Cutoffs for fuzzy matching
# PERMISSIVE_CUTOFF: For general typo correction (e.g., BCS405A vs BCS405A.)
# STRICT_CUTOFF: For potentially confusing codes (e.g., BCS405 vs BCS406)
PERMISSIVE_CUTOFF = 0.86 
STRICT_CUTOFF = 0.99      

# ==============================================
# ARGUMENT PARSING
# ==============================================
# Sets to store subject codes for filtering
FILTERED_SUBJECT_CODES_NORM = set()  # Normalized codes for filtering
SUBJECT_CODES_EXPLICIT = set()  # Canonical codes provided by the user (for headers)

# Parse subject codes from command line argument (if provided)
# Subject codes are passed as a JSON string array
if len(sys.argv) > 1:
    try:
        # The subject codes list is passed as a JSON string (sys.argv[1])
        subject_codes_list = json.loads(sys.argv[1])
        
        # Store normalized codes for both filtering and explicit header list
        for code in subject_codes_list:
             norm_code = normalize_code(code)
             FILTERED_SUBJECT_CODES_NORM.add(norm_code)
             SUBJECT_CODES_EXPLICIT.add(norm_code)  # Use normalized code for strict headers
             
        print(f"📝 Filtering results for {len(FILTERED_SUBJECT_CODES_NORM)} subjects.")
    except json.JSONDecodeError:
        print("⚠️ Failed to parse subject codes argument. Processing all subjects found.")
    except Exception as e:
        print(f"⚠️ Error processing command-line arguments: {e}. Processing all subjects found.")

# ==============================================
# STEP 0: Setup Excel File 
# ==============================================
# Load existing Excel file or create a new one
# If file exists, we'll update it with new data; otherwise, create from scratch
if os.path.exists(EXCEL_OUTPUT_FILENAME):
    print(f"🔄 Loading existing Excel file: {EXCEL_OUTPUT_FILENAME}")
    wb = load_workbook(EXCEL_OUTPUT_FILENAME)
    ws = wb.active
else:
    print(f"➕ Creating new Excel file: {EXCEL_OUTPUT_FILENAME}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Set up initial header for USN column (merged across 2 rows)
    header_font = Font(bold=True)
    ws.cell(row=1, column=1).value = "University Seat Number"
    ws.cell(row=1, column=1).font = header_font
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")


# ==============================================
# STEP 1: Helper Functions
# ==============================================

def get_existing_subject_columns(ws):
    """
    Get a mapping of existing subject codes to their column indices in the Excel sheet.
    
    Subject columns are spaced 4 columns apart (Internal, External, Total, Result).
    Column 1 is reserved for USN.
    
    Args:
        ws: Excel worksheet object
        
    Returns:
        Dictionary mapping subject code to column index
    """
    subject_cols = {}
    # Iterate through columns starting from 2, stepping by 4 (each subject takes 4 columns)
    for col in range(2, ws.max_column + 1, 4):
        code = ws.cell(row=1, column=col).value
        if code:
            subject_cols[code] = col
    return subject_cols

def get_next_empty_row(ws, usn):
    """
    Find the row for a USN. If USN already exists, return its row; otherwise, find next empty row.
    
    This allows updating existing USN data or adding new USNs to the sheet.
    
    Args:
        ws: Excel worksheet object
        usn: University Seat Number to find or add
        
    Returns:
        Row index for the USN
    """
    # Create a map of existing USNs to their row numbers (starting from row 3, after headers)
    existing_usn_map = {ws.cell(row=i, column=1).value: i for i in range(3, ws.max_row + 2)}

    # If USN already exists, return its row
    if usn in existing_usn_map:
        return existing_usn_map[usn]
    
    # Find the next empty row in column A (USN column)
    for row in range(3, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    # If no empty row found, return the next row after the last row
    return ws.max_row + 1

def ensure_and_sort_subject_headers(ws, new_codes):
    """
    Ensure all subject codes have columns in the Excel sheet, and sort all headers alphabetically.
    
    This function:
    1. Merges new codes with existing codes
    2. Sorts all codes alphabetically
    3. Clears and recreates headers with proper formatting
    4. Each subject gets 4 columns: Internal, External, Total, Result
    
    Args:
        ws: Excel worksheet object
        new_codes: Set of new subject codes to add
    """
    header_font = Font(bold=True)
    
    # Get existing subject codes from the sheet
    existing_codes = list(get_existing_subject_columns(ws).keys())
    all_codes = existing_codes[:]
    
    # Add new codes that don't already exist
    for code in new_codes:
        if code and code not in all_codes:
            all_codes.append(code)

    # Sort all codes alphabetically for consistent ordering
    all_codes.sort()

    # Clear old headers by unmerging merged cells
    merged_to_unmerge = [str(m) for m in ws.merged_cells.ranges if m.min_row <= 2]
    for m in merged_to_unmerge:
        try: 
            ws.unmerge_cells(m)
        except ValueError: 
            pass  # Cell might not be merged

    # Clear existing header values (rows 1 and 2)
    if ws.max_column > 1:
        for col in range(2, ws.max_column + 1):
            ws.cell(row=1, column=col).value = None
            ws.cell(row=2, column=col).value = None

    # Recreate headers with proper formatting
    for i, code in enumerate(all_codes):
        # Calculate starting column (each subject takes 4 columns)
        col = 2 + (i * 4)
        
        # Merge cells in row 1 for subject code (spans 4 columns)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        
        # Set subject code in merged cell
        cell1 = ws.cell(row=1, column=col)
        cell1.value = code
        cell1.alignment = Alignment(horizontal="center")
        cell1.font = header_font
        
        # Set sub-headers in row 2: Internal, External, Total, Result
        for j, name in enumerate(["Internal", "External", "Total", "Result"]):
            cell2 = ws.cell(row=2, column=col + j)
            cell2.value = name
            cell2.alignment = Alignment(horizontal="center")
            cell2.font = header_font

def matches_filter(code, filter_set):
    """
    Check if a code matches any code in the filter set, considering last letter matching.
    
    This function handles cases where:
    - User provides "BCS405" and JSON has "BCS405A", "BCS405B", etc. -> should match
    - User provides "BCS405A" and JSON has "BCS405" -> should match
    - User provides "BCS405A" and JSON has "BCS405B" -> should NOT match (different last letters)
    
    Args:
        code: Normalized code to check
        filter_set: Set of normalized filter codes
        
    Returns:
        True if code matches any filter code, False otherwise
    """
    if not code or not filter_set:
        return False
    
    norm_code = normalize_code(code)
    if not norm_code:
        return False
    
    # 1. Exact match
    if norm_code in filter_set:
        return True
    
    # 2. Last letter matching: If code ends with a letter, check if base code is in filter
    if len(norm_code) > 1 and norm_code[-1].isalpha() and norm_code[-1].isupper():
        base_code = norm_code[:-1]  # Remove last letter
        if base_code in filter_set:
            return True
    
    # 3. Reverse check: If any filter code ends with a letter, check if its base matches this code
    for filter_code in filter_set:
        if len(filter_code) > 1 and filter_code[-1].isalpha() and filter_code[-1].isupper():
            filter_base = filter_code[:-1]  # Remove last letter from filter code
            if filter_base == norm_code:
                return True
    
    return False


def canonicalize_code(raw_code, existing_codes):
    """
    Canonicalize subject code by matching it to existing codes or creating a new one.
    
    This function uses a hybrid approach:
    - Strict matching for potentially confusing codes (e.g., BCS405 vs BCS406)
    - Permissive matching for general typo correction (e.g., BCS405A vs BCS405A.)
    - Last letter matching: If code ends with a letter and base code exists, use base code
      (e.g., BCS405A, BCS405B, BCS405C all map to BCS405 if BCS405 exists)
    
    Args:
        raw_code: Raw subject code extracted from JSON
        existing_codes: Set of existing canonical codes to match against
        
    Returns:
        Canonical subject code (matched or new)
    """
    norm = normalize_code(raw_code)
    if not norm: return raw_code

    existing_norm_map = {normalize_code(c): c for c in existing_codes}

    # 1. Exact normalized match
    if norm in existing_norm_map:
        return existing_norm_map[norm]

    # 2. Last letter matching: If code ends with a letter (A-Z), check if base code exists
    # This handles cases like BCS405A, BCS405B, BCS405C -> BCS405
    if len(norm) > 1 and norm[-1].isalpha() and norm[-1].isupper():
        base_code = norm[:-1]  # Remove last letter
        # Check if base code exists in normalized existing codes
        if base_code in existing_norm_map:
            return existing_norm_map[base_code]

    # Determine cutoff: Strict for codes matching the pattern of confusing VTU codes
    # VTU codes typically follow pattern: [BCV][A-Z]{3}\d{3} (e.g., BCS405, BCS406)
    # For these, use strict matching to avoid confusing similar codes
    if 6 <= len(norm) <= 7 and re.match(r"[BCV][A-Z]{3}\d{3}", norm):
        cutoff = STRICT_CUTOFF  # 0.99 for high stability (avoid false matches)
    else:
        cutoff = PERMISSIVE_CUTOFF  # 0.86 for general typo correction (e.g., BCS405A vs BCS405A.)

    # 3. Fuzzy match: Find similar codes using string similarity
    close = difflib.get_close_matches(norm, list(existing_norm_map.keys()), n=1, cutoff=cutoff) 

    if close:
        # Found a similar code, return the canonical version
        return existing_norm_map[close[0]]
    else:
        # 4. New code: No match found, return normalized code as new canonical code
        return norm

# Define color fills for Excel cells based on result
# Red: Failed (F), Green: Passed (P), Yellow: Absent (A)
fill_red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
fill_green = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
fill_yellow = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")


# ==============================================
# STEP 2: Main Processing Loop - Read JSON Files
# ==============================================

# Get all unique existing headers from Excel sheet for canonicalization
existing_headers = set(get_existing_subject_columns(ws).keys())

# Data structures to collect processed data
all_canonical_codes_found = set()  # All unique canonical subject codes found
data_to_write = {}  # Nested dictionary: {USN: {CanonicalCode: {Internal: 40, External: 49, Total: 89, Result: "P"}}}

print("🔎 Starting JSON file iteration and data collection...")

# Find all JSON files in the input folder (files ending with '_gemini_output.json')
json_files = [f for f in os.listdir(JSON_INPUT_FOLDER) if f.endswith('_gemini_output.json')]

# Check if any JSON files exist
if not json_files:
    print(f"⚠️ No JSON files found in '{JSON_INPUT_FOLDER}'. Exiting.")
    # Clean up empty folder
    if os.path.exists(JSON_INPUT_FOLDER):
        shutil.rmtree(JSON_INPUT_FOLDER)
        print(f"🗑️ Deleted empty folder: {JSON_INPUT_FOLDER}")
    exit()

# Process each JSON file
for filename in json_files:
    json_path = os.path.join(JSON_INPUT_FOLDER, filename)
    
    print(f"   Processing file: {filename}")
    
    # Read JSON file
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            json_output = json.load(f)
    except Exception as e:
        print(f"❌ Error reading or parsing JSON file {filename}: {e}. Skipping.")
        continue

    # Extract USN from JSON (required field)
    usn = (json_output.get("USN") or "NOTFOUND").strip().upper()
    if usn == "NOTFOUND":
        print(f"⚠️ USN not found in {filename}. Skipping.")
        continue

    # Ensure we don't process the same USN multiple times (handle duplicate files)
    if usn in data_to_write:
        print(f"⚠️ USN {usn} already processed. Skipping duplicate file {filename}.")
        continue

    # Initialize data structure for this USN
    data_to_write[usn] = {}
    
    # Process each subject in the JSON file
    for sub in json_output.get("Subjects", []):
        raw_code = sub.get("Code", "").strip()
        if not raw_code: 
            continue  # Skip subjects without codes

        # Canonicalize the extracted code (match to existing or create new)
        # Use existing headers, user-provided codes, and newly found codes for matching
        # This ensures that if user provides "BCS405", codes like "BCS405A" will be canonicalized to "BCS405"
        current_canonical_set = existing_headers.union(SUBJECT_CODES_EXPLICIT).union(all_canonical_codes_found)
        canon_code = canonicalize_code(raw_code, current_canonical_set) 
        
        # Filtering logic: Only process subjects in the filter list (if filter is set)
        # Use matches_filter to handle last letter matching (e.g., BCS405 matches BCS405A, BCS405B, etc.)
        if FILTERED_SUBJECT_CODES_NORM and not matches_filter(canon_code, FILTERED_SUBJECT_CODES_NORM):
            # Also check the raw code before canonicalization in case canonicalization changed it
            if not matches_filter(raw_code, FILTERED_SUBJECT_CODES_NORM):
                print(f"   Skipping subject {raw_code} for {usn}: Not in filter list.")
                continue

        # Add to set of found canonical codes
        all_canonical_codes_found.add(canon_code)
        
        # Extract marks data
        internal = sub.get("Internal")
        external = sub.get("External")
        total = sub.get("Total")
        result = sub.get("Result", "").strip().upper()

        # Calculate total if not provided (sum of internal and external)
        if (total is None or total == "") and isinstance(internal, int) and isinstance(external, int):
            total = internal + external
            
        # Store data for this USN and subject
        data_to_write[usn][canon_code] = {
            "Internal": internal,
            "External": external,
            "Total": total,
            "Result": result,
        }

print(f"✅ Collected data for {len(data_to_write)} unique USNs.")

# ==============================================
# STEP 3: Write Data to Excel
# ==============================================

# Step 3.1: Update headers with all collected canonical codes
print("✍️ Updating Excel headers...")

# Header logic: Determine which subject codes should appear as columns
# The headers should be the union of:
# 1. The explicit codes passed by the user (SUBJECT_CODES_EXPLICIT)
# 2. Any existing headers that match the filter (to preserve old data)
# 3. Any canonical codes found in this run (all_canonical_codes_found)

headers_to_use = set()

if FILTERED_SUBJECT_CODES_NORM:
    # Filtering is enabled: only include subjects in the filter list
    # 1. Add all canonical codes found in the JSONs (which are already filtered)
    headers_to_use.update(all_canonical_codes_found)
    
    # 2. Add all existing Excel headers that match the filter (to preserve existing data)
    # Use matches_filter to handle last letter matching (e.g., BCS405 matches BCS405A)
    # Also canonicalize existing headers to base codes if they match
    for h in existing_headers:
        norm_h = normalize_code(h)
        if matches_filter(h, FILTERED_SUBJECT_CODES_NORM):
            # Canonicalize the existing header using user-provided codes and found codes
            # This ensures BCS405A becomes BCS405 if BCS405 is in the filter
            canon_h = canonicalize_code(h, SUBJECT_CODES_EXPLICIT.union(all_canonical_codes_found))
            headers_to_use.add(canon_h)
    
    # 3. Add the explicitly passed normalized codes as headers
    # This ensures all expected columns exist, even if no data was found for them
    headers_to_use.update(SUBJECT_CODES_EXPLICIT)
    
else:
    # No filtering: include all codes found and all existing codes
    headers_to_use = all_canonical_codes_found.union(existing_headers)
    
# Update Excel headers with the determined subject codes
ensure_and_sort_subject_headers(ws, headers_to_use)
# Get the mapping of subject codes to column indices after header update
subject_cols = get_existing_subject_columns(ws)

# Step 3.2: Iterate through collected data and write to Excel sheet
print("✍️ Writing marks to Excel sheet...")
for usn, subjects in data_to_write.items():
    
    # Find the correct row for this USN (existing row or new row)
    row = get_next_empty_row(ws, usn)
    # Write USN in column 1 (or overwrite if updating existing row)
    ws.cell(row=row, column=1).value = usn

    # Write marks for each subject
    for code, info in subjects.items():
        if code in subject_cols:
            col = subject_cols[code]  # Get starting column for this subject
            
            # Clear old data in the 4 cells for this subject (Internal, External, Total, Result)
            # This ensures we don't have stale data when updating
            for offset in range(4):
                 ws.cell(row=row, column=col + offset).value = None
                 ws.cell(row=row, column=col + offset).fill = PatternFill(fill_type=None)  # Clear color fill

            # Write marks: Internal, External, Total (columns 0, 1, 2)
            for i, key in enumerate(["Internal", "External", "Total"]):
                value = info.get(key, "")
                # Try to convert to integer, otherwise keep as-is
                try:
                    cell_value = int(value)
                except (ValueError, TypeError):
                    cell_value = value
                
                ws.cell(row=row, column=col + i).value = cell_value
                
            # Write Result and apply color coding (column 3)
            result_cell = ws.cell(row=row, column=col + 3)
            result = info.get("Result", "").upper()
            result_cell.value = result
            
            # Apply color fill based on result: Red (F), Green (P), Yellow (A)
            if result == "F":
                result_cell.fill = fill_red
            elif result == "P":
                result_cell.fill = fill_green
            elif result == "A":
                result_cell.fill = fill_yellow

# Step 3.3: Save the Excel file
try:
    wb.save(EXCEL_OUTPUT_FILENAME)
    print(f"✅ Successfully processed all JSON files and saved data to: {EXCEL_OUTPUT_FILENAME}")
except Exception as e:
    print(f"❌ ERROR: Could not save Excel file. Please ensure '{EXCEL_OUTPUT_FILENAME}' is closed. Error: {e}")


# ==============================================
# STEP 4: CLEANUP
# ==============================================
# Delete the temporary JSON folder after processing
# This helps keep the workspace clean and reduces disk usage
if os.path.exists(JSON_INPUT_FOLDER):
    try:
        shutil.rmtree(JSON_INPUT_FOLDER)
        print(f"🗑️ Successfully deleted temporary folder: {JSON_INPUT_FOLDER}")
    except OSError as e:
        print(f"❌ ERROR: Could not delete folder {JSON_INPUT_FOLDER}. Check permissions or file locks. Error: {e}")